##This program creates the report from the input excel file.

##Importing the required libraries
from tkinter import ttk
from tkinter.ttk import Progressbar, Treeview
from tkinter import *
from PIL import Image,ImageTk
import os
import datetime
from PIL import Image,ImageTk
from tkinter.filedialog import askopenfile
from tkinter import messagebox
from openpyxl.reader.excel import load_workbook
import pandas as pd
import numpy as np
import threading
import random
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

path = os.path.dirname(os.path.realpath(__file__))

w=Tk()


width_of_window = 427
height_of_window = 250
screen_width = w.winfo_screenwidth()
screen_height = w.winfo_screenheight()
x_coordinate = (screen_width/2)-(width_of_window/2)
y_coordinate = (screen_height/2)-(height_of_window/2)
w.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))


w.overrideredirect(1)


s = ttk.Style()
s.theme_use('clam')
s.configure("red.Horizontal.TProgressbar", foreground='red', background='#dafbff')
progress=Progressbar(w,style="red.Horizontal.TProgressbar",orient=HORIZONTAL,length=500,mode='determinate',)

#############progressbar          33333333333333333333333333333
def new_win():
  # w.destroy()
    q=Tk()
    #q.iconbitmap(path+"/icon.ico")
    q.title('Report Extractor Tool')
    q.iconbitmap(path+"\\assets\\ret_icon.ico")
    #q.geometry('427x250')

    width_of_mwindow = 600
    height_of_mwindow = 400
    screen_mwidth = q.winfo_screenwidth()
    screen_mheight = q.winfo_screenheight()
    x_mcoordinate = (screen_mwidth/2)-(width_of_mwindow/2)
    y_mcoordinate = (screen_mheight/2)-(height_of_mwindow/2)
    q.geometry("%dx%d+%d+%d" %(width_of_mwindow,height_of_mwindow,x_mcoordinate,y_mcoordinate))
    q.resizable(0,0)

    my_notebook=ttk.Notebook(q)
    my_notebook.pack(fill="both", expand=1)

#hourly report tab_"每小时报告"选项卡
    tab1=Frame(my_notebook,width=600, height=400, bg="white")
    tab1.pack(fill="both", expand=1)

    img_tab1= ImageTk.PhotoImage(Image.open(path+"\\images\\hourly_tab.png"))
    img_lab=Label(tab1, image=img_tab1, width=150, height=150, borderwidth=0)
    img_lab.place(x=40, y=70)

    tab1_l2=Label(tab1,text='Welcome to,',fg='black',bg='white',font="Calibri 11")
    tab1_l2.place(x=205, y=80)

    tab1_l3=Label(tab1,text='REPORT',fg='black',bg="white", font='Calibri 18 bold')
    tab1_l3.place(x=205,y=100)

    tab1_l4=Label(tab1,text='EXTRACTOR',fg='black',bg="white", font='Calibri 18' )
    tab1_l4.place(x=295,y=100)

    tab1_l5=Label(tab1,text='TOOL',fg='black',bg="white", font='Calibri 18 bold')
    tab1_l5.place(x=205,y=130)

    tab1_l6=Label(tab1,text=' ',fg='black',bg="white", font="Calibri 6 ")
    tab1_l6.place(x=20, y=350)
    tab1_l7=Label(tab1, text='', fg='black',bg="white", font="Calibri 6 ")
    tab1_l7.place(x=480, y=350)
    rand1=random.randint(2,4)

    tab1_l8=Label(tab1,text='Total Shippments:',fg='black',bg='white',font="Calibri 8")
    tab1_l8.place(x=205, y=210)
    tab1_l9=Label(tab1,text='000000',fg='black',bg='white',font="Calibri 8")
    tab1_l9.place(x=295, y=210)
    tab1_l10=Label(tab1,text='Total DBO:',fg='black',bg='white',font="Calibri 8")
    tab1_l10.place(x=205, y=235)
    tab1_l11=Label(tab1,text='000000',fg='black',bg='white',font="Calibri 8")
    tab1_l11.place(x=295, y=235)
    tab1_l12=Label(tab1,text='Total Exception:',fg='black',bg='white',font="Calibri 8")
    tab1_l12.place(x=205, y=260)
    tab1_l13=Label(tab1,text='000000',fg='black',bg='white',font="Calibri 8")
    tab1_l13.place(x=295, y=260)
    tab1_l14=Label(tab1,text='Total Timeout:',fg='black',bg='white',font="Calibri 8")
    tab1_l14.place(x=205, y=285)
    tab1_l15=Label(tab1,text='000000',fg='black',bg='white',font="Calibri 8")
    tab1_l15.place(x=295, y=285)

    tab1.pack(fill="both",expand=1)
    
    tab1_frame=Frame(tab1, width=605, height=40, bg=a)
    tab1_frame.pack()

    #img2_tab1= ImageTk.PhotoImage(Image.open(path+"/RET_up.png"))
    #img2_lab=Label(tab1_frame, image=img2_tab1, width=18, height=18, borderwidth=0)
    #img2_lab.place(x=560, y=10)

    #tip_tab1=Balloon(tab1)
    #tip_tab1.bind_widget(img2_lab,balloonmsg="What's new.!Now you're able to know about the TIMEOUT values.")

    tab1_l1=Label(tab1,text='Select the file to be sorted:',fg='white',bg=a,font="Calibri 8 ")
    tab1_l1.place(x=40, y=10)

    tab1_b2= Button(tab1, text="Import File", font="Calibri 8 bold",command=lambda:threading.Thread(target=open_file).start(), width=15, bg="white", fg=a, borderwidth=0, relief=GROOVE)
    tab1_b2.place(x=180,y=10)



##Opens the file
    def open_file():
        file=askopenfile(parent=tab1, mode='rb', title="Choose the file to be execuited", filetypes=[("Excel File","*.xlsx"), ("Excel File","*.xls")])
        file_name_path = str(file)[26:-2]
        file_name_time=datetime.datetime.now().strftime('%Y-%M-%D %H-%M%S')
        tab1_l6.config(text="Extracting Data")
        tab1_l7.config(text=f"Generating data in {rand1} minutes")
        tab1_l9.config(text=" ")
        tab1_l11.config(text=" ")
        tab1_l13.config(text=" ")
        tab1_l15.config(text=" ")
        if file:
            if (driver_fn(file_name_path)):
                messagebox.showinfo("Information", "The Report has been generated")
                tab1_l6.config(text="Data Generated")
                tab1_l7.config(text=" ")    
        else:
            messagebox.showerror("Alert", "No File Uploaded") 
            tab1_l6.config(text=" ")
            tab1_l7.config(text=" ")
            tab1_l9.config(text="00000")
            tab1_l11.config(text="00000")
            tab1_l13.config(text="00000")
            tab1_l15.config(text="00000")
##Driver function
    def driver_fn(file_path):
        shipments_per_hour_per_space_name_list = read_excel(file_path)
        write_excel(file_path, shipments_per_hour_per_space_name_list)
        return True


##Reads excel
    def read_excel(xl_input):
        shipments_per_hour_list = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        shipments_per_hour_per_space_name_list = [[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]]
        total_dbo_count = 0
        total_exception_count = 0
        time_out_gt_10 = 0
        time_out_count=0
        df_xl_read = pd.read_excel(xl_input, dtype={'Begin Time':str, 'End Time':str})#, sheet_name=0)
        begin_time_list = df_xl_read['Begin Time'].tolist()
        space_name_list = df_xl_read['Space Name'].tolist()
        chute_id_list = df_xl_read['Chute Id'].tolist()
        end_time_list = df_xl_read['End Time'].tolist()
    
        for x in range(0, len(begin_time_list)):
            begin_date_time_details = begin_time_list[x].strip().split(" ")
            space_name = space_name_list[x].strip()
            for i in range(0, 24):
                if (i < 10):
                    i_start_time = "0" + str(i) + ":00:00.000"
                    i_end_time = "0" + str(i) + ":59:59.999"
                else:
                    i_start_time = str(i) + ":00:00.000"
                    i_end_time = str(i) + ":59:59.999"

                if (begin_date_time_details[1] >= i_start_time and begin_date_time_details[1] <= i_end_time):
                    shipments_per_hour_list[i] += 1
                
                    for j in range(1,37):
                        if (j < 10):
                            j_space_name = "SP00" + str(j)
                        else:
                            j_space_name = "SP0"+ str(j)

                        if (space_name == j_space_name):
                            shipments_per_hour_per_space_name_list[i][j-1] = shipments_per_hour_per_space_name_list[i][j-1] + 1
                            break

            if ("DBO" in chute_id_list[x]):
                total_dbo_count += 1
                check_float = isinstance(end_time_list[x], float)
                if (check_float == False):
                    end_date_time_details = end_time_list[x].strip().split(" ")
                    time_diff = time_out_calc(begin_date_time_details[1], end_date_time_details[1], begin_date_time_details[0], end_date_time_details[0])
                    if (time_diff >= 1000.000):
                        time_out_gt_10 += 1
            tab1_l11.config(text=total_dbo_count)

            if ("EXCEPTION" in chute_id_list[x]):
                total_exception_count += 1
            tab1_l13.config(text=total_exception_count)

            if ("TIMEOUT" in chute_id_list[x]):
                time_out_count+=1
            #print(time_out_count)
            tab1_l15.config(text=time_out_count)
            tab1_l9.config(text=len(df_xl_read))

        return(shipments_per_hour_per_space_name_list)





    def time_out_calc(begin_time, end_time, begin_date, end_date):
        begin_time = float(begin_time.replace(":", ""))
        end_time = float(end_time.replace(":", ""))
        if (begin_date == end_date):
            return end_time - begin_time
        else:
            end_time += 240000.000
            return end_time - begin_time


##Creates the output excel
    def write_excel(file_path, shipments_per_hour_per_space_name_list):
        shipments_per_hour_per_space_name_list_transpose = [[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]]
        
        shipments_per_hour_per_space_name_list_formatted = []

        for i in range(6, 24):
            shipments_per_hour_per_space_name_list_formatted.append(shipments_per_hour_per_space_name_list[i])
        for i in range(0, 6): 
            shipments_per_hour_per_space_name_list_formatted.append(shipments_per_hour_per_space_name_list[i])

        for i in range(0, 36):
            for j in range(0, 24):
                shipments_per_hour_per_space_name_list_transpose[i][j] = shipments_per_hour_per_space_name_list_formatted[j][i]

        df_to_write = pd.DataFrame(shipments_per_hour_per_space_name_list_transpose)
        sp_col = []
        for i in range(1, 37):
            str1 = "SP0"
            if (i < 10):
                str1 = "SP00"
            str1 = str1 + str(i)
            sp_col.append(str1)
        
        str1 = ""
        tod_rows = []
        for i in range(6, 23):
            str1 = str(i) + ":00 to  " + str(i+1) + ":00"
            tod_rows.append(str1)
        tod_rows.append("23:00 to 0:00")
        for i in range(0, 6):
            str1 = str(i) + ":00 to " + str(i+1) + ":00"
            tod_rows.append(str1)
            
        df_to_write_2 = pd.DataFrame(sp_col)
        df_pick_up_port_number_heading = pd.DataFrame(["Pick-up port Number"])
        df_time_heading = pd.DataFrame(["Time"])
        df_to_write.columns = tod_rows
        df_to_write.index = sp_col
        with pd.ExcelWriter((os.path.dirname(file_path)+"/Hourly_Report.xlsx")) as writer:
            df_to_write.to_excel(writer, sheet_name='Hourly count of pick up ports', na_rep='', startrow=1, startcol=1, merge_cells=True, inf_rep='inf', verbose=True)



##Applies the formatting to the generated report
    def apply_formatting(workbook_name):
        workbook=load_workbook(filename = workbook_name)
        workbook_active_sheet = workbook.active
        cell_1 = workbook_active_sheet.cell(row = 6, column = 3)
        cell_1.value = "Pick-up port Number"
        cell_1.font  = Font(name="Arial", b=True, color="FF0000")
        workbook_active_sheet.merge_cells('D6:AA6')
        cell_2 = workbook_active_sheet.cell(row = 6, column = 4)
        cell_2.value = "Time"
        cell_2.alignment = Alignment(horizontal="center", vertical="center")
        cell_2.font  = Font(name="Arial", b=True, color="FF0000")
        workbook.save(workbook_name)    



#minutely throughput tab_分钟吞吐量选项卡
    tab2=Frame(my_notebook, width=600, height= 400, bg="white")
    tab2.pack(fill="both", expand=1 )

    tab2_frame=Frame(tab2, width=605, height=40, bg=a)
    tab2_frame.pack()
    tab2_l4=Label(tab2,text='Select the file to be sorted:',fg='black',bg="white",font="Calibri 8 ")
    tab2_l4.place(x=20, y=50)
    tab2_b2= Button(tab2, text="Import File", command=lambda:threading.Thread(target=open_file_mint).start(), font="Calibri 8 bold", width=15, bg=a, fg="white",borderwidth=0)
    tab2_b2.place(x=160,y=50)

    tab2_l1=Label(tab2,text='Enter the desired maximum value:  ',fg='white', bg=a, font="Calibri 8 ")
    tab2_l1.place(x=20, y=10)
    global tab2_e1
    a_var = StringVar()
    tab2_e1=Entry(tab2, width=15,textvariable=a_var, relief=SUNKEN)
    tab2_e1.place(x=190,y=10)
    tab2_b1=Button(tab2, text="Reset", command=lambda:res(),font="Calibri 7 bold", width=7, bg="#ff5252", fg="white", borderwidth=0)
    tab2_b1.place(x=290,y=12)

    #img_tab2= ImageTk.PhotoImage(Image.open("C:\\Users\\Jacob Thomas\\Documents\\Python Program\\My Programs\\mintgraph.png"))
    #img_lab=Label(tab2, image=img_tab2, width=600, height=200)
    #img_lab.place(x=0, y=140)
    img_tab2= ImageTk.PhotoImage(Image.open(path+"\\images\\minut_tab.png"))
    img2_lab=Label(tab2, image=img_tab2, width=250, height=250, borderwidth=0)
    img2_lab.place(x=340, y=90)

    style=ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview", background="white", foreground="black", rowheight=20, fieldbackground="white")
    style.map('Treeview', background=[('selected', a)])
    tab2_tree=ttk.Treeview(tab2)
    tab2_tree['columns']= ('Minutely Throughput', 'Count')
    tab2_tree.column('#0', width=50, minwidth=10)
    tab2_tree.column('Minutely Throughput',anchor=CENTER, width=150)
    tab2_tree.column('Count', anchor=CENTER, width=100)
    tab2_tree.heading('#0', text="Sl. No.", anchor=CENTER)
    tab2_tree.heading('Minutely Throughput', text="Minutely Throughput", anchor=CENTER)
    tab2_tree.heading('Count', text="Count", anchor=CENTER)
    tab2_tree.place(x=20, y=110)

    

    tab2_l2=Label(tab2,text='The Maximum value calculated:',fg='black', bg="white", font="Calibri 8")
    tab2_l2.place(x=20, y=80)
    global tab2_l3
    tab2_l3=Label(tab2,text=' 000 ',fg='black', bg="white", font="Calibri 8 bold")
    tab2_l3.place(x=175, y=80)
    tab2_l5=Label(tab2,text=' ',fg='black',bg="white", font="Calibri 6 ")
    tab2_l5.place(x=20, y=350)
    tab2_l6=Label(tab2, text='', fg='black',bg="white", font="Calibri 6 ")
    tab2_l6.place(x=480, y=350)
    rand=random.randint(2,5)

    def res():
        tab2_e1.delete(0, 'end')
        tab2_l3.config(text="000")
        for rec in tab2_tree.get_children():
             tab2_tree.delete(rec)

    def open_file_mint():
         v = int(a_var.get())
         file=askopenfile(parent=tab2, mode='rb', title="Choose the file to be execuited", filetypes=[("Excel File","*.xlsx"), ("Excel File","*.xls")])
         file_name_path = str(file)[26:-2]
         tab2_l5.config(text="Loading...")
         tab2_l3.config(text="000")
         tab2_l6.config(text=f"Generating data in {rand} minutes")
         for rec in tab2_tree.get_children():
             tab2_tree.delete(rec)
         #tab2_l5 =Label.config(file) 
         #tab2_e1.config(state='disabled')
         if file:
             data=pd.read_excel(file)
             tab2_l5.config(text="Loading...")
             #converting column data to date time format
             data["Begin Time"]=pd.to_datetime(data["Begin Time"])
             #extracting hour minute data from column_从列中提取小时分钟数据
             begin_time= data["Begin Time"].dt.strftime("%H:%M")
             #hour minute list
             mnt=[ ]
             mng=[ ]
             #minutely count list
             mntc=[ ]
             mngc=[ ]
             start_time=""
             #extracting data
             for i in range (0, 24):
                 for j in range (0,60):
                      if (i < 10):
                         if(j < 10): 
                             start_time = "0" + str(i) + ":0"+str(j)

                         else:
                             start_time = "0" + str(i) + ":"+str(j)
                      else:
                         if(j < 10): 
                             start_time = str(i) + ":0"+str(j)
                         else:
                             start_time = str(i) + ":"+str(j)
                      mnt.append(start_time)
                      ctr=0
                      for x in begin_time:
                          if start_time in str(x):
                              ctr += 1
                             #print(ctr)                                      
                      mntc.append(ctr)
                      if ctr >= v:
                         print("Minutely Throughput at ",start_time,": ", ctr)
                         mng.append(start_time)
                         mngc.append(ctr)
             max_value = None
             for num in mntc:
                 if (max_value is None or num > max_value):
                     max_value = num

             if (len(mngc) ==0):
                 messagebox.showerror("Alert", f"No data found above {v} shippments per minute. Maximum value found is {max_value}.\n \nTry again with a lesser value.")
                 tab2_l5.config(text=" ")
                 tab2_l6.config(text=" ")

             else:
                 x=np.array(mng)
                 y=np.array(mngc)
                 max_value = None
                 for num in mntc:
                     if (max_value is None or num > max_value):
                         max_value = num
                 #print("The highest minutely throughput",max_value)
                 tab2_l3.config(text=max_value)
                 for h in range(len(mngc)):
                     tab2_tree.insert('', h, values=(mng[h], mngc[h]))
             #fig = go.Figure(data=go.Bar(x=mng, y=mngc, text=y, textposition='outside', width=0.4))
             #fig.update_layout(title_text=f"Minutely Throughput - Above {v} Shipments",title_x=0.5,title_font_size=18)
             #fig.update_layout(title_text="Minutely Throughput - Above 250 Shipments",title_x=0.5,title_font_size=15)
             #fig.write_html('minutely.html', auto_open=True)
                 with open((os.path.dirname(str(file_name_path))+"/minthrptgh.txt"), 'w') as f:
                     for dm, dc in zip(mng,mngc):
                         f.writelines('Minutely Throughput at ')
                         f.writelines(str(dm))
                         f.writelines(': ')
                         f.writelines(str(dc))
                         f.writelines('\n')
                      #f.writeslines("{}\t{}\n".format(mnt,mntc))
                 f.close
                 tab2_l5.config(text="Done")
                 tab2_l6.config(text=" ")
                 messagebox.showinfo("Info", "Data generated")
         else:
             tab2_l5.config(text=" ")
             tab2_l6.config(text=" ")
             messagebox.showerror("Alert", "No File Uploaded")    

    #chute report
    tab3=Frame(my_notebook, width=600, height= 400, bg="white")
    tab3.pack(fill="both", expand=1 )
    tab3_frame=Frame(tab3, width=605, height=40, bg=a)
    tab3_frame.pack()

    tab3_l1=Label(tab3_frame,text='Select the file to be sorted:',fg='white',bg=a,font="Calibri 8 ")
    tab3_l1.place(x=20, y=10)
    tab3_b1= Button(tab3_frame, text="Import File", command=lambda:threading.Thread(target=open_file_chuter).start(), font="Calibri 8 bold",borderwidth=0, width=15, bg="white", fg=a, relief=GROOVE)
    tab3_b1.place(x=160,y=10)

    tab3_l2=Label(tab3, text='', fg='black',bg="white", font="Calibri 6 ")
    tab3_l2.place(x=480, y=350)

    style=ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview", background="white", foreground="black", rowheight=20, fieldbackground="white")
    style.map('Treeview', background=[('selected', a)])
    tab3_tree=ttk.Treeview(tab3)
    tab3_tree['columns']= ('Chute Name', 'Count')
    tab3_tree.column('#0', width=50, minwidth=10)
    tab3_tree.column('Chute Name',anchor=CENTER, width=150)
    tab3_tree.column('Count', anchor=CENTER, width=100)
    tab3_tree.heading('#0', text="Sl. No.", anchor=CENTER)
    tab3_tree.heading('Chute Name', text="Chute Name", anchor=CENTER)
    tab3_tree.heading('Count', text="Count", anchor=CENTER)
    tab3_tree.place(x=270, y=90)

    img_tab3= ImageTk.PhotoImage(Image.open(path+"\\images\\chute_tab.png"))
    img2_lab=Label(tab3, image=img_tab3, width=250, height=250, borderwidth=0)
    img2_lab.place(x=13, y=70)

    def open_file_chuter():
         file=askopenfile(parent=tab3, mode='rb', title="Choose the file to be execuited", filetypes=[("Excel File","*.xlsx"), ("Excel File","*.xls")])
         file_name_path = str(file)[26:-2]
         tab3_l2.config(text=f"Generating data in {rand} minutes")
         if file:
             data=pd.read_excel(file)
             chute_data=data['Chute Id']
             chute_name=""
             chnm=[ ]
             chcn=[ ]
             #chcdx=np.array([84, 87, 90, 93, 96, 99, 102, 105, 108, 111, 114, 117, 120, 123, 126, 129,132, 135, 138, 141, 144, 147, 150, 153, 156, 159, 162, 165, 168, 171, 174, 177, 180, 183 ])
             #chcdy=np.array([86, 89, 92, 95, 98, 101, 104, 107, 110, 113, 116, 119, 122, 125, 128, 131, 134, 137, 140, 143, 146, 149, 152, 155, 158, 161, 164, 167, 170, 173, 176, 179, 182, 185])

             for i in range(1,519):
                 if (i<10):
                     chute_name="CH"+"0"+"0"+str(i)
                     #print(chute_name)
              
                 else:
                     if(i<100):
                         chute_name="CH"+"0"+str(i)
                          #print(chute_name)
                    
                     else:
                          chute_name="CH"+str(i)
                          #print(chute_name)
                          
                 chnm.append(chute_name)       
                 cn=0
                 for j in chute_data:
                     if chute_name in str(j):
                         cn +=1
                 chcn.append(cn)
                 #print(cn)

             print(chnm)
             print(chcn)

             for c in range(len(chnm)):
                 tab3_tree.insert('', c, values=(chnm[c], chcn[c]))

             with open((os.path.dirname(str(file_name_path))+"/Chute_conc.txt"), 'w') as f:
                 for dm, dc in zip(chnm,chcn):
                     f.writelines(str(dm))
                     f.writelines('- ')
                     f.writelines(str(dc))
                     f.writelines('\n')
                     #f.writeslines("{}\t{}\n".format(mnt,mntc))
             f.close
             tab3_l2.config(text="")
             messagebox.showinfo("Info", "Chute Data generated. Please check the folder.")

    #minutley report
    tab4=Frame(my_notebook, width=600, height= 400, bg="white")
    tab4.pack(fill="both", expand=1 )
    tab4_frame=Frame(tab4, width=605, height=40, bg=a)
    tab4_frame.pack()

    tab4_l1=Label(tab4_frame,text='Select the file to be sorted:',fg='white',bg=a,font="Calibri 8 ")
    tab4_l1.place(x=20, y=10)
    tab4_b1= Button(tab4_frame, text="Import File", command=lambda:threading.Thread(target=open_file_mintr).start(), font="Calibri 8 bold", borderwidth=0, width=15, bg="white", fg=a, relief=GROOVE)
    tab4_b1.place(x=160,y=10)

    tab4_l2=Label(tab4,text='The Maximum value calculated:',fg='black', bg="white", font="Calibri 8")
    tab4_l2.place(x=20, y=50)
    tab4_l3=Label(tab4,text=' 000 ',fg='black', bg="white", font="Calibri 8 bold")
    tab4_l3.place(x=175, y=50)
    tab4_l4=Label(tab4,text='Count of time more than 230/min (from 230 to 239):',fg='black', bg="white", font="Calibri 8")
    tab4_l4.place(x=280, y=85)
    tab4_l5=Label(tab4,text='',fg='black', bg="white", font="Calibri 8 bold")
    tab4_l5.place(x=525, y=85)
    tab4_l6=Label(tab4,text='Count of time more than 240/min (from 240 to 249):',fg='black', bg="white", font="Calibri 8")
    tab4_l6.place(x=280, y=105)
    tab4_l7=Label(tab4,text='',fg='black', bg="white", font="Calibri 8 bold")
    tab4_l7.place(x=525, y=105)
    tab4_l8=Label(tab4,text='Count of time more than 250/min (more than 250):',fg='black', bg="white", font="Calibri 8")
    tab4_l8.place(x=280, y=125)
    tab4_l9=Label(tab4,text='',fg='black', bg="white", font="Calibri 8 bold")
    tab4_l9.place(x=525, y=125)

    tab4_tree=Treeview(tab4)
    tab4_tree['columns']= ('Time', 'Count')
    tab4_tree.column('#0', width=80, minwidth=50)
    tab4_tree.column('Time',anchor=W, width=80)
    tab4_tree.column('Count', anchor=CENTER, width=80)
    tab4_tree.heading('#0', text="Sl. No.", anchor=W)
    tab4_tree.heading('Time', text="Time", anchor=W)
    tab4_tree.heading('Count', text="Count", anchor=CENTER)
    tab4_tree.place(x=20,y=85)
    tab4_l10=Label(tab4, text='', fg='black',bg="white", font="Calibri 6 ")
    tab4_l10.place(x=480, y=350)
    #tab4_tree_scroll=Scrollbar(tab4_tree)
    #tab4_tree_scroll.pack(side=RIGHT, fill=Y)

    def open_file_mintr():
         file=askopenfile(parent=tab4, mode='rb', title="Choose the file to be execuited", filetypes=[("Excel File","*.xlsx"), ("Excel File","*.xls")])
         file_name_path = str(file)[26:-2]
         for rec in tab4_tree.get_children():
             tab4_tree.delete(rec)
         tab4_l10.config(text="Data is generating...")
         tab4_l5.config(text="")
         tab4_l7.config(text="")
         tab4_l9.config(text="")
         tab4_l3.config(text="")
         if file:
             data=pd.read_excel(file)
               #converting column data to date time format
             data["Begin Time"]=pd.to_datetime(data["Begin Time"])
             #extracting hour minute data from column
             begin_time= data["Begin Time"].dt.strftime("%H:%M")
             #hour minute list
             mnt=[ ]
             mng=[ ]
             #minutely count list
             mntc=[ ]
             mngc=[ ]
             start_time=""
             #extracting data
             for i in range (0, 24):
                 for j in range (0,60):
                      if (i < 10):
                         if(j < 10): 
                             start_time = "0" + str(i) + ":0"+str(j)
                         else:
                             start_time = "0" + str(i) + ":"+str(j)
                      else:
                         if(j < 10): 
                             start_time = str(i) + ":0"+str(j)
                         else:
                             start_time = str(i) + ":"+str(j)
                      mnt.append(start_time)
                      ctr=0
                      for x in begin_time:
                          if start_time in str(x):
                              ctr += 1
                             #print(ctr)
                      mntc.append(ctr)
                      #if ctr >= a:
                         #print("Minutely Throughput at ",start_time,": ", ctr)
                         #mng.append(start_time)
                         #mngc.append(ctr) 
             x=np.array(mng)
             y=np.array(mngc)
             max_value = None
             for num in mntc:
                 if (max_value is None or num > max_value):
                     max_value = num
             #print("The highest minutely throughput",max_value)
             tab4_l3.config(text=max_value)
             c1=0
             c2=0
             c3=0
             for mc in mntc:
                 if (mc>=230 and mc<240):
                     c1+=1
                 elif (mc>=240 and mc<250):
                     c2+=1
                 elif(mc>=250):
                     c3+=1
             print("Count of time more than 230",c1,'\n')
             print("Count of time more than 240",c2,'\n')
             print("Count of time more than 250",c3,'\n')
             for x in range(len(mnt)):
                  tab4_tree.insert('', x, values=(mnt[x], mntc[x]))
             #fig = go.Figure(data=go.Bar(x=mng, y=mngc, text=y, textposition='outside', width=0.4))
             #fig.update_layout(title_text=f"Minutely Throughput - Above {a} Shipments",title_x=0.5,title_font_size=15)
             #fig.update_layout(title_text="Minutely Throughput - Above 250 Shipments",title_x=0.5,title_font_size=15)
             #fig.write_html('minutely.html', auto_open=True)
             with open((os.path.dirname(str(file_name_path))+"/minthrptdt.txt"), 'w') as f:
                  for dm, dc in zip(mnt,mntc):
                      f.writelines('Minutely Data at ')
                      f.writelines(str(dm))
                      f.writelines(': ')
                      f.writelines(str(dc))
                      f.writelines('\n')
                      #f.writeslines("{}\t{}\n".format(mnt,mntc))
             f.close
             tab4_l5.config(text=c1)
             tab4_l7.config(text=c2)
             tab4_l9.config(text=c3)
             tab4_l10.config(text="Data is generated")
             messagebox.showinfo("Info", "Data generated")
         else:
             messagebox.showerror("Alert", "No File Uploaded")
             tab4_l10.config(text=" ") 

    my_notebook.add(tab1,text= "Hourly Report")
    my_notebook.add(tab2,text= "Minutely Throughput")
    my_notebook.add(tab3,text= "Chute Report")
    my_notebook.add(tab4,text= "Minutely Report")

    #f1=Frame(q, width=300, height=400, bg=a)
    #f1.place(x=0,y=0)
    #browse_text= StringVar()
    #browse_btn= Button(q, textvariable=browse_text, command=lambda:open_file(), font="Calibri 8 bold", width=45, height=3, bg=a, fg="white")
    #browse_text.set("Browse file")
    #browse_btn.place(x=250, y=250)

    q.mainloop()


def bar():

    l4=Label(w,text='Loading...',fg='white',bg=a)
    lst4=('Calibri (Body)',10)
    l4.config(font=lst4)
    l4.place(x=18,y=210)
    
    import time
    r=0
    for i in range(100):
        progress['value']=r
        w.update_idletasks()
        time.sleep(0.03)
        r=r+1
    
    w.destroy()
    new_win()
        
    
progress.place(x=-10,y=235)




#############
# frame 333333333333333333333333
#
###########





'''

def rgb(r):
    return "#%02x%02x%02x" % r
#Frame(w,width=432,height=241,bg=rgb((100,100,100))).
'''
a='#ff8045'
Frame(w,width=427,height=241,bg=a).place(x=0,y=0)  #249794
b1=Button(w,width=10,height=1,text='Get Started',command=bar,border=0,fg=a,bg='white')
b1.place(x=170,y=200)


######## Label

l1=Label(w,text='REPORT',fg='white',bg=a)
lst1=('Calibri (Body)',18,'bold')
l1.config(font=lst1)
l1.place(x=50,y=80)

l2=Label(w,text='EXTRACTOR',fg='white',bg=a)
lst2=('Calibri (Body)',18)
l2.config(font=lst2)
l2.place(x=155,y=82)

l3=Label(w,text='TOOL',fg='white',bg=a)
lst3=('Calibri (Body)',15)
l3.config(font=lst3)
l3.place(x=50,y=110)

l4=Label(w,text='V 2.1',fg='white',bg=a)
lst4=('Calibri (Body)',10)
l4.config(font=lst4)
l4.place(x=50,y=150)

  
# l4=Label(w,text=' Programmed by Jacob Thomas | HIKVISION INDIA',fg='white',bg=a)
# lst4=('Calibri (Body)',4)
# l4.config(font=lst4)
# l4.place(x=300,y=20)

w.mainloop()



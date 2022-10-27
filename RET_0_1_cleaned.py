from tkinter import *
from PIL import Image,ImageTk
import time
from tkinter.filedialog import askopenfile
from tkinter import ttk
from tkinter import messagebox
import os
import pandas as pd
import openpyxl
import datetime
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

path = os.path.dirname(os.path.realpath(__file__))

#Frontend
root=Tk()
root.resizable(0,0)

Width=600
Height=400

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width/2) - (Width/2)
y = (screen_height/2) - (Height/2)
root.geometry("+%d+%d" % (x, y))

canvas=Canvas(root, width=Width, height=Height, bg="#383838")
canvas.grid(columnspan=3,rowspan=5)

#Background Image- Update
image=ImageTk.PhotoImage(Image.open(path+"/img.png"))
canvas.create_image(0,0, anchor=NW, image=image)


#instructions
instructions= Label(root, text="Select the file to be sorted", font="Calibri 22 bold", fg="#545454", bg="#EBEBEB")
instructions.grid(column=1,row=0)

#copyrights
copyrights= Label(root, text="Developed by Prama Hikvision India - Robotics & Factory Automation Team.", font="Calibri 6 ", fg="#545454", bg="#EBEBEB")
copyrights.grid(column=1,row=4)

#Application Title & Icon  
root.title("Report Extractor Tool")
root.iconbitmap(path+"/RETicon.ico")

def open_file():
    browse_text.set("Loading...")
    file=askopenfile(parent=root, mode='rb', title="Choose the file to be execuited", filetypes=[("Excel File","*.xlsx"), ("Excel File","*.xls")])
    file_name_path = str(file)[26:-2]
    browse_text.set("Browse file")
    if file:
        my_progress=ttk.Progressbar(root, orient=HORIZONTAL, length=200, mode='indeterminate')
        my_progress.grid(column=1, row=3, pady=20)
        my_progress.start(10)
        
        if (driver_fn(file_name_path)):
            messagebox.showinfo("Information", "The Report has been generated") 
            
    else:
        messagebox.showerror("Alert", "No File Uploaded") 

def driver_fn(file_path):
     shipments_per_hour_per_space_name_list = read_excel(file_path)
     write_excel(file_path, shipments_per_hour_per_space_name_list)
     return True

def read_excel(xl_input):
    shipments_per_hour_list = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    shipments_per_hour_per_space_name_list = [[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]]
    total_dbo_count = 0
    total_exception_count = 0
    time_out_gt_10 = 0
    df_xl_read = pd.read_excel(xl_input)#, sheet_name=0)
    begin_time_list = df_xl_read['Begin Time'].tolist()
    space_name_list = df_xl_read['Space Name'].tolist()
    chute_id_list = df_xl_read['Chute Id'].tolist()
    end_time_list = df_xl_read['End Time'].tolist()
   
    for x in range(0, len(begin_time_list)):
        begin_date_time_details = begin_time_list[x].strip().split(" ")
        space_name = space_name_list[x].strip()
        for i in range(0, 24):
            if (i < 10):
                i_start_time = "0" + str(i) + ":00:00.000"
                i_end_time = "0" + str(i) + ":59:59.999"
            else:
                i_start_time = str(i) + ":00:00.000"
                i_end_time = str(i) + ":59:59.999"

            if (begin_date_time_details[1] >= i_start_time and begin_date_time_details[1] <= i_end_time):
                shipments_per_hour_list[i] += 1
            
                for j in range(1,37):
                    if (j < 10):
                        j_space_name = "SP00" + str(j)
                    else:
                        j_space_name = "SP0"+ str(j)

                    if (space_name == j_space_name):
                        shipments_per_hour_per_space_name_list[i][j-1] = shipments_per_hour_per_space_name_list[i][j-1] + 1
                        break

        if ("DBO" in chute_id_list[x]):
            total_dbo_count += 1
            check_float = isinstance(end_time_list[x], float)
            if (check_float == False):
                end_date_time_details = end_time_list[x].strip().split(" ")
                time_diff = time_out_calc(begin_date_time_details[1], end_date_time_details[1], begin_date_time_details[0], end_date_time_details[0])
                if (time_diff >= 1000.000):
                    time_out_gt_10 += 1

        if ("EXCEPTION" in chute_id_list[x]):
            total_exception_count += 1
    return(shipments_per_hour_per_space_name_list)

def time_out_calc(begin_time, end_time, begin_date, end_date):
    begin_time = float(begin_time.replace(":", ""))
    end_time = float(end_time.replace(":", ""))
    if (begin_date == end_date):
        return end_time - begin_time
    else:
        end_time += 240000.000
        return end_time - begin_time

def write_excel(file_path, shipments_per_hour_per_space_name_list):
    shipments_per_hour_per_space_name_list_transpose = [[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 
0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]]
    
    shipments_per_hour_per_space_name_list_formatted = []

    for i in range(6, 24):
        shipments_per_hour_per_space_name_list_formatted.append(shipments_per_hour_per_space_name_list[i])
    for i in range(0, 6): 
        shipments_per_hour_per_space_name_list_formatted.append(shipments_per_hour_per_space_name_list[i])

    for i in range(0, 36):
        for j in range(0, 24):
            shipments_per_hour_per_space_name_list_transpose[i][j] = shipments_per_hour_per_space_name_list_formatted[j][i]

    df_to_write = pd.DataFrame(shipments_per_hour_per_space_name_list_transpose)
    sp_col = []
    for i in range(1, 37):
        str1 = "SP0"
        if (i < 10):
            str1 = "SP00"
        str1 = str1 + str(i)
        sp_col.append(str1)
    
    str1 = ""
    tod_rows = []
    for i in range(6, 23):
        str1 = str(i) + ":00 to  " + str(i+1) + ":00"
        tod_rows.append(str1)
    tod_rows.append("23:00 to 0:00")
    for i in range(0, 6):
        str1 = str(i) + ":00 to " + str(i+1) + ":00"
        tod_rows.append(str1)
        
    df_to_write_2 = pd.DataFrame(sp_col)
    df_pick_up_port_number_heading = pd.DataFrame(["Pick-up port Number"])
    df_time_heading = pd.DataFrame(["Time"])
    df_to_write.columns = tod_rows
    df_to_write.index = sp_col
    with pd.ExcelWriter((os.path.dirname(file_path)+"/Report.xlsx")) as writer:
        df_to_write.to_excel(writer, sheet_name='Hourly count of pick up ports', na_rep='', startrow=2, startcol=2, merge_cells=True, inf_rep='inf', verbose=True)

def apply_formatting(workbook_name):
    workbook = load_workbook(filename = workbook_name)
    workbook_active_sheet = workbook.active
    cell_1 = workbook_active_sheet.cell(row = 6, column = 3)
    cell_1.value = "Pick-up port Number"
    cell_1.font  = Font(name="Arial", b=True, color="FF0000")
    workbook_active_sheet.merge_cells('D6:AA6')
    cell_2 = workbook_active_sheet.cell(row = 6, column = 4)
    cell_2.value = "Time"
    cell_2.alignment = Alignment(horizontal="center", vertical="center")
    cell_2.font  = Font(name="Arial", b=True, color="FF0000")
    workbook.save(workbook_name)    

#browse
browse_text= StringVar()
browse_btn= Button(root, textvariable=browse_text, command=lambda:open_file(), font="Calibri 8 bold", width=50, height=3, bg="#de7d00", fg="white")
browse_text.set("Browse file")
browse_btn.grid(column=1, row=1)

root.mainloop()
